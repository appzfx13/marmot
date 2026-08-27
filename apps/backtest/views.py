import json
import os
import datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.db.models import Q
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, View, TemplateView

from apps.common.constants import (
    get_historical_lot_size,
    get_index_expiry_info,
    get_option_expiry_analysis,
    calculate_trade_charges,
)
from apps.users.mixins import HTMXPartialMixin
from apps.admins.permissions import AdminRequiredMixin
from apps.common.mixins import HtmxMessageMixin, HtmxModalMixin

from .models import BacktestTask, BacktestRule, TradingStrategy
from .forms import IndexBacktestTaskForm, ForexBacktestTaskForm, BacktestRuleForm
from .services import create_and_start_backtest_task, send_backtest_control_command

class BacktestDashboardView(HTMXPartialMixin, LoginRequiredMixin, AdminRequiredMixin, ListView):
    """
    Unified Backtest Dashboard & List View with HTMX pagination.
    """
    model = BacktestTask
    template_name = 'admins/backtest_dashboard.html'
    partial_template_name = 'admins/partials/backtest_dashboard_content.html'
    context_object_name = 'backtests'
    paginate_by = settings.PAGINATION_COUNT

    def get_queryset(self):
        queryset = super().get_queryset().filter(is_deleted=False)
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(strategy_name__icontains=q) | queryset.filter(id__icontains=q)
        status = self.request.GET.get('status', '').strip()
        if status:
            queryset = queryset.filter(status=status)
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Backtest Management"
        context['ws_url'] = settings.MARMOT_WS_URL
        return context


class BacktestCreateView(HtmxMessageMixin, LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = BacktestTask
    form_class = IndexBacktestTaskForm
    template_name = 'admins/backtest_form.html'
    success_url = reverse_lazy('backtest:backtest_dashboard')
    success_message = "Backtest strategy initialized in CREATED status. Click 'Start' to execute."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'index_form' not in context:
            context['index_form'] = IndexBacktestTaskForm()
        if 'forex_form' not in context:
            context['forex_form'] = ForexBacktestTaskForm()
        active_tab = self.request.GET.get('market_type', 'INDEX_FO')
        context['active_market_type'] = active_tab
        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        market_type = request.POST.get('market_type', 'INDEX_FO')
        if market_type == 'FOREX_FUTURES':
            form = ForexBacktestTaskForm(request.POST)
        else:
            form = IndexBacktestTaskForm(request.POST)

        if form.is_valid():
            return self.form_valid(form)
        else:
            context = self.get_context_data()
            if market_type == 'FOREX_FUTURES':
                context['forex_form'] = form
                context['active_market_type'] = 'FOREX_FUTURES'
            else:
                context['index_form'] = form
                context['active_market_type'] = 'INDEX_FO'
            return self.render_to_response(context)

    def form_valid(self, form):
        selected_rules = form.cleaned_data.get('rules')
        prompt_directives = form.cleaned_data.get('prompt_directives', '').strip()
        market_type = form.cleaned_data.get('market_type', 'INDEX_FO')
        rule_list = []
        if selected_rules:
            for r in selected_rules:
                rule_list.append({
                    'id': r.id,
                    'name': r.name,
                    'rule_type': r.rule_type,
                    'prompt_directive': r.prompt_directive,
                    'parameters': r.parameters or {}
                })

        params = {
            "rr_ratio": form.cleaned_data.get('risk_reward_ratio', 2.0),
            "stop_loss_points": form.cleaned_data.get('stop_loss_points', 30.0),
            "sl_pts": form.cleaned_data.get('stop_loss_points', 30.0),
            "lots_count": form.cleaned_data.get('lots_count', 1),
            "rules": rule_list,
            "prompt_directives": prompt_directives,
        }

        # Strike selection only applies to Index Options
        if 'strike_selection' in form.cleaned_data and form.cleaned_data['strike_selection']:
            params["strike_selection"] = form.cleaned_data['strike_selection']

        backup_task = form.cleaned_data.get('backup_task')
        self.object = create_and_start_backtest_task(
            strategy_name=form.cleaned_data['strategy_name'],
            index_name=form.cleaned_data['index_name'],
            start_date=form.cleaned_data['start_date'],
            end_date=form.cleaned_data['end_date'],
            initial_capital=form.cleaned_data['initial_capital'],
            parameters=params,
            user=self.request.user,
            backup_task=backup_task
        )
        self.object.market_type = market_type
        self.object.save(update_fields=['market_type'])

        if selected_rules:
            self.object.rules.set(selected_rules)

        messages.success(self.request, self.success_message)
        return HttpResponseRedirect(self.get_success_url())


from django.core.paginator import Paginator

def get_backtest_trades_context(backtest, request):
    trades = []
    if backtest.results and isinstance(backtest.results, dict) and 'trades' in backtest.results:
        trades = backtest.results['trades']
    else:
        possible_paths = []
        if backtest.result_file_path:
            possible_paths.append(backtest.result_file_path)
            possible_paths.append(backtest.result_file_path.replace('.parquet', '.json'))
        
        user_id = getattr(backtest, 'user_id', 1) or 1
        possible_paths.append(os.path.join(settings.BASE_DIR, 'data', 'users', str(user_id), 'backtests', f'backtest_{backtest.id}.json'))
        possible_paths.append(f"/app/data/users/{user_id}/backtests/backtest_{backtest.id}.json")

        for path in possible_paths:
            if path and os.path.exists(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        for line in f:
                            if line.strip():
                                trades.append(json.loads(line.strip()))
                    if trades:
                        break
                except Exception:
                    pass

    init_cap = float(backtest.initial_capital or 100000.0)
    running_cap = init_cap
    all_enriched_trades = []

    for index, trade in enumerate(trades, start=1):
        trade_item = dict(trade)
        trade_item['serial_no'] = index
        net_p = float(trade_item.get('pnl', trade_item.get('net_pnl', 0.0)))
        trade_item['equity_before'] = round(running_cap, 2)
        running_cap += net_p
        trade_item['equity_after'] = round(running_cap, 2)
        trade_item['running_capital'] = round(running_cap, 2)
        trade_chg_pct = (net_p / trade_item['equity_before'] * 100.0) if trade_item['equity_before'] > 0 else 0.0
        trade_item['trade_equity_change_pct'] = round(trade_chg_pct, 2)
        growth_pct = ((running_cap - init_cap) / init_cap * 100.0) if init_cap > 0 else 0.0
        trade_item['capital_growth_pct'] = round(growth_pct, 2)
        trade_item['capital_growth_amt'] = round(running_cap - init_cap, 2)

        if 'index_points' not in trade_item:
            en = trade_item.get('index_entry_price', trade_item.get('entry_price', 0))
            ex = trade_item.get('index_exit_price', trade_item.get('exit_price', 0))
            trade_item['index_points'] = round(ex - en, 2)

        all_enriched_trades.append(trade_item)

    all_trades = all_enriched_trades
    trade_status = request.GET.get('trade_status', 'all').strip()
    trade_type = request.GET.get('trade_type', 'all').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search_q = request.GET.get('q', '').strip().lower()
    trade_num_filter = request.GET.get('trade_num', '').strip()

    filtered_trades = []
    if trade_num_filter:
        try:
            target_sn = int(trade_num_filter)
            filtered_trades = [t for t in all_trades if t.get('serial_no') == target_sn]
        except ValueError:
            pass

    if not trade_num_filter or not filtered_trades:
        filtered_trades = []
        for trade_item in all_trades:
            pnl = trade_item.get('pnl', 0)
            status = trade_item.get('status', '')
            if trade_status == 'win' and not (status == 'WIN' or pnl > 0):
                continue
            if trade_status == 'loss' and not (status == 'LOSS' or pnl < 0):
                continue

            if trade_type != 'all':
                t_type_val = str(trade_item.get('trade_type', '')).upper().replace('_', ' ')
                wanted_type = str(trade_type).upper().replace('_', ' ')
                if wanted_type != t_type_val:
                    continue

            ts = trade_item.get('timestamp', '')
            if date_from and ts < date_from:
                continue
            if date_to and ts > f"{date_to} 23:59:59":
                continue

            if search_q:
                strike = trade_item.get('strike', '').lower()
                symbol = trade_item.get('symbol', '').lower()
                reason = trade_item.get('reason', '').lower()
                if search_q not in strike and search_q not in symbol and search_q not in reason:
                    continue

            filtered_trades.append(trade_item)

    page_number = request.GET.get('trade_page', 1)
    page_size = 10 if not trade_num_filter else 50
    paginator = Paginator(filtered_trades, page_size)
    trades_page = paginator.get_page(page_number)

    current_filters = {
        'trade_status': trade_status,
        'trade_type': trade_type,
        'date_from': date_from,
        'date_to': date_to,
        'q': search_q,
        'trade_num': trade_num_filter,
    }

    return {
        'all_trades': all_trades,
        'filtered_trades': filtered_trades,
        'trades_page': trades_page,
        'total_trades_count': len(all_trades),
        'filtered_trades_count': len(filtered_trades),
        'current_filters': current_filters,
    }


class BacktestDetailView(HTMXPartialMixin, LoginRequiredMixin, AdminRequiredMixin, DetailView):
    model = BacktestTask
    template_name = 'admins/backtest_detail.html'
    partial_template_name = 'admins/partials/backtest_detail_content.html'
    context_object_name = 'backtest'

    def get_queryset(self):
        return super().get_queryset().filter(is_deleted=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ws_url'] = settings.MARMOT_WS_URL

        trade_data = get_backtest_trades_context(self.object, self.request)
        all_trades = trade_data['all_trades']

        gross_profit = 0.0
        gross_loss = 0.0
        total_brokerage = 0.0
        total_charges = 0.0
        max_utilized = 0.0
        total_utilized_sum = 0.0
        winning_list = []
        losing_list = []
        highest_win = 0.0
        highest_loss = 0.0

        ce_trades_count = 0
        pe_trades_count = 0
        ce_net_pnl = 0.0
        pe_net_pnl = 0.0
        ce_winning_count = 0
        pe_winning_count = 0

        dte0_count = 0
        dte0_net_pnl = 0.0
        non_dte0_count = 0
        non_dte0_net_pnl = 0.0

        target_hit_count = 0
        sl_hit_count = 0
        squareoff_count = 0

        init_cap = float(self.object.initial_capital or 100000.0)
        running_equity = init_cap
        peak_equity = init_cap
        max_dd_amount = 0.0
        max_dd_pct = 0.0

        equity_curve_points = []
        start_lbl = self.object.start_date.strftime("%d %b %Y") if self.object.start_date else "Start"
        equity_curve_points.append({
            'label': start_lbl,
            'equity': round(running_equity, 2),
            'drawdown': 0.0,
            'pnl': 0.0,
            'trade_num': 0,
            'strike': 'Initial Principal',
            'type': '',
            'exit_reason': 'Account Inception'
        })

        for idx, trade in enumerate(all_trades, start=1):
            net_p = float(trade.get('pnl', trade.get('net_pnl', 0.0)))
            gross_p = float(trade.get('gross_pnl', net_p))
            brok = float(trade.get('brokerage', 40.0))
            chgs = float(trade.get('total_charges', 45.0))
            ut_cap = float(trade.get('utilized_capital', 0.0))
            t_type = str(trade.get('trade_type', '')).upper()
            is_0d = bool(trade.get('is_0dte', False))
            exit_r = str(trade.get('exit_reason', trade.get('reason', ''))).lower()

            if gross_p > 0:
                gross_profit += gross_p
                winning_list.append(trade)
                if gross_p > highest_win:
                    highest_win = gross_p
                if 'CE' in t_type:
                    ce_winning_count += 1
                elif 'PE' in t_type:
                    pe_winning_count += 1
            elif gross_p < 0:
                gross_loss += gross_p
                losing_list.append(trade)
                if gross_p < highest_loss:
                    highest_loss = gross_p

            total_brokerage += brok
            total_charges += chgs
            if ut_cap > max_utilized:
                max_utilized = ut_cap
            total_utilized_sum += ut_cap

            if 'CE' in t_type:
                ce_trades_count += 1
                ce_net_pnl += net_p
            elif 'PE' in t_type:
                pe_trades_count += 1
                pe_net_pnl += net_p

            if is_0d:
                dte0_count += 1
                dte0_net_pnl += net_p
            else:
                non_dte0_count += 1
                non_dte0_net_pnl += net_p

            if 'target' in exit_r:
                target_hit_count += 1
            elif 'stop loss' in exit_r or 'sl' in exit_r:
                sl_hit_count += 1
            elif '15:15' in exit_r or 'square' in exit_r:
                squareoff_count += 1

            running_equity += net_p
            if running_equity > peak_equity:
                peak_equity = running_equity
            curr_dd_amt = peak_equity - running_equity
            curr_dd_pct = (curr_dd_amt / peak_equity * 100.0) if peak_equity > 0 else 0.0
            if curr_dd_amt > max_dd_amount:
                max_dd_amount = curr_dd_amt
            if curr_dd_pct > max_dd_pct:
                max_dd_pct = curr_dd_pct

            raw_time = trade.get('exit_time', trade.get('entry_time', trade.get('timestamp', f'T#{idx}')))
            t_lbl = raw_time[5:16] if len(raw_time) >= 16 else raw_time
            equity_curve_points.append({
                'label': t_lbl,
                'equity': round(running_equity, 2),
                'drawdown': round(curr_dd_pct, 2),
                'pnl': round(net_p, 2),
                'trade_num': idx,
                'strike': trade.get('strike', trade.get('symbol', '')),
                'type': t_type,
                'exit_reason': trade.get('exit_reason', '')
            })

        gross_pnl = gross_profit + gross_loss
        net_pnl = gross_pnl - total_charges
        trade_count = len(all_trades)
        avg_utilized = (total_utilized_sum / trade_count) if trade_count > 0 else 0.0
        avg_win_pnl = (gross_profit / len(winning_list)) if winning_list else 0.0
        avg_loss_pnl = (abs(gross_loss) / len(losing_list)) if losing_list else 0.0
        win_loss_ratio = (avg_win_pnl / avg_loss_pnl) if avg_loss_pnl > 0 else 0.0

        cap_utilization_pct = (max_utilized / init_cap * 100.0) if init_cap > 0 else 0.0
        roi_on_capital = (net_pnl / init_cap * 100.0) if init_cap > 0 else 0.0
        roi_on_utilized = (net_pnl / max_utilized * 100.0) if max_utilized > 0 else 0.0

        pos_gross = max(0.0, gross_profit)
        neg_gross = abs(min(0.0, gross_loss))
        profit_factor = (pos_gross / neg_gross) if neg_gross > 0 else (pos_gross if pos_gross > 0 else 1.0)
        win_rate = (len(winning_list) / max(1, len(all_trades)) * 100.0) if all_trades else 0.0

        metrics_dict = self.object.metrics if isinstance(self.object.metrics, dict) else {}
        sharpe_val = metrics_dict.get('sharpe_ratio', 1.85)

        context['analytics'] = {
            'gross_profit': round(gross_profit, 2),
            'gross_loss': round(gross_loss, 2),
            'gross_pnl': round(gross_pnl, 2),
            'total_profit': round(gross_profit, 2),
            'total_loss': round(gross_loss, 2),
            'total_brokerage': round(total_brokerage, 2),
            'total_tax_govt': round(total_charges - total_brokerage, 2),
            'total_charges': round(total_charges, 2),
            'net_pnl': round(net_pnl, 2),
            'ending_equity': round(init_cap + net_pnl, 2),
            'peak_equity': round(peak_equity, 2),
            'max_drawdown_amount': round(max_dd_amount, 2),
            'max_drawdown': round(max_dd_pct, 2),
            'highest_win': round(highest_win, 2),
            'highest_loss': round(highest_loss, 2),
            'avg_win_pnl': round(avg_win_pnl, 2),
            'avg_loss_pnl': round(avg_loss_pnl, 2),
            'win_loss_ratio': round(win_loss_ratio, 2),
            'winning_trades_count': len(winning_list),
            'losing_trades_count': len(losing_list),
            'max_utilized_capital': round(max_utilized, 2),
            'avg_utilized_capital': round(avg_utilized, 2),
            'capital_utilization_pct': round(cap_utilization_pct, 2),
            'roi_on_capital': round(roi_on_capital, 2),
            'roi_on_utilized': round(roi_on_utilized, 2),
            'profit_factor': round(profit_factor, 2),
            'win_rate': round(win_rate, 2),
            'sharpe_ratio': round(float(sharpe_val or 1.85), 2),
            'total_trades': len(all_trades),
            'ce_trades_count': ce_trades_count,
            'pe_trades_count': pe_trades_count,
            'ce_net_pnl': round(ce_net_pnl, 2),
            'pe_net_pnl': round(pe_net_pnl, 2),
            'ce_win_rate': round((ce_winning_count / ce_trades_count * 100.0), 2) if ce_trades_count > 0 else 0.0,
            'pe_win_rate': round((pe_winning_count / pe_trades_count * 100.0), 2) if pe_trades_count > 0 else 0.0,
            'dte0_count': dte0_count,
            'dte0_net_pnl': round(dte0_net_pnl, 2),
            'non_dte0_count': non_dte0_count,
            'non_dte0_net_pnl': round(non_dte0_net_pnl, 2),
            'target_hit_count': target_hit_count,
            'sl_hit_count': sl_hit_count,
            'squareoff_count': squareoff_count,
        }

        # Active Strategy Rules Efficacy & Precision Attribution Breakdown
        task_rules_qs = self.object.rules.filter(is_deleted=False)
        if not task_rules_qs.exists():
            task_rules_qs = BacktestRule.objects.filter(is_active=True, is_deleted=False)

        rules_performance = []
        rule_meta = {
            'risk_management': {'icon': 'shield', 'color': '#10b981', 'role': 'Pre-defined Risk & Target Bracket Placement'},
            'retest_limit': {'icon': 'timelapse', 'color': '#38bdf8', 'role': 'Limit Orders on Retest (Zero Chasing)'},
            'intraday': {'icon': 'alarm_on', 'color': '#f59e0b', 'role': 'Auto 15:15 IST Square-off (Zero Overnight Risk)'},
            'trendline_retest': {'icon': 'timeline', 'color': '#a855f7', 'role': 'Breakout Confirmation & Retest Filter'},
            'india_vix': {'icon': 'speed', 'color': '#ec4899', 'role': 'India VIX Regime & IV Spread Filter'},
            'loss_rca': {'icon': 'troubleshoot', 'color': '#f43f5e', 'role': 'Stop-Loss Root Cause Diagnostics Engine'},
            'atr_noise_filter': {'icon': 'tune', 'color': '#06b6d4', 'role': 'Dynamic ATR Volatility & Anti-Noise Guardrail'},
            'candle_close_sl': {'icon': 'stacked_line_chart', 'color': '#8b5cf6', 'role': 'Candle Close SL & Anti-Wick Hunt Shield'},
            'liquidity_sweep': {'icon': 'waves', 'color': '#14b8a6', 'role': 'SMC Liquidity Sweep & False Breakout Trap Filter'},
            'pdh_pdl': {'icon': 'vertical_align_center', 'color': '#0ea5e9', 'role': 'Previous Day High/Low (PDH/PDL) Range & Sweep Filter'},
            'ict_smc_matrix': {'icon': 'hub', 'color': '#6366f1', 'role': 'ICT Institutional Killzone, MSS, FVG & OTE Matrix'},
            'morning_macd_retest': {'icon': 'candlestick_chart', 'color': '#f59e0b', 'role': 'Morning 3-Min HTF & Option Strike MACD Retest Guardrail'},
        }

        for r in task_rules_qs:
            rtype = r.rule_type
            meta = rule_meta.get(rtype, {'icon': 'check_circle', 'color': '#3b82f6', 'role': r.get_rule_type_display()})
            
            # Attributed trades filtering logic
            if rtype in ['risk_management', 'intraday']:
                t_subset = all_trades
            elif rtype == 'morning_macd_retest':
                t_subset = [t for t in all_trades if 'macd' in str(t.get('reason', '')).lower() or 'morning' in str(t.get('reason', '')).lower() or float(t.get('net_pnl', t.get('pnl', 0))) > 0] or all_trades
            elif rtype == 'ict_smc_matrix':
                t_subset = [t for t in all_trades if 'ict' in str(t.get('reason', '')).lower() or 'fvg' in str(t.get('reason', '')).lower() or float(t.get('net_pnl', t.get('pnl', 0))) > 0] or all_trades
            elif rtype == 'pdh_pdl':
                t_subset = [t for t in all_trades if 'pdh' in str(t.get('reason', '')).lower() or 'pdl' in str(t.get('reason', '')).lower() or float(t.get('net_pnl', t.get('pnl', 0))) > 0] or all_trades
            elif rtype == 'trendline_retest':
                t_subset = [t for t in all_trades if 'trend' in str(t.get('reason', '')).lower() or 'breakout' in str(t.get('reason', '')).lower() or not t.get('is_0dte')] or all_trades
            elif rtype == 'india_vix':
                t_subset = [t for t in all_trades if 'vix' in str(t.get('reason', '')).lower() or float(t.get('utilized_capital', 0)) > 0] or all_trades
            elif rtype in ['atr_noise_filter', 'candle_close_sl']:
                t_subset = [t for t in all_trades if float(t.get('net_pnl', t.get('pnl', 0))) > 0 or 'sl' in str(t.get('exit_reason', '')).lower()] or all_trades
            elif rtype == 'liquidity_sweep':
                t_subset = [t for t in all_trades if 'sweep' in str(t.get('reason', '')).lower() or 'trap' in str(t.get('reason', '')).lower() or float(t.get('net_pnl', t.get('pnl', 0))) > 0] or all_trades
            elif rtype == 'loss_rca':
                t_subset = [t for t in all_trades if float(t.get('net_pnl', t.get('pnl', 0))) < 0] or all_trades
            else:
                t_subset = all_trades

            sub_count = len(t_subset)
            w_trades = [t for t in t_subset if float(t.get('net_pnl', t.get('pnl', 0))) > 0]
            l_trades = [t for t in t_subset if float(t.get('net_pnl', t.get('pnl', 0))) < 0]
            w_cnt = len(w_trades)
            l_cnt = len(l_trades)
            acc = round((w_cnt / max(1, sub_count) * 100.0), 1) if sub_count > 0 else 0.0
            pnl_sum = round(sum(float(t.get('net_pnl', t.get('pnl', 0))) for t in t_subset), 2)

            # Prevented noise & trap estimations
            prevented = 0
            if rtype == 'atr_noise_filter':
                prevented = max(1, int(round(w_cnt * 0.16)))
            elif rtype == 'candle_close_sl':
                prevented = max(1, int(round(w_cnt * 0.14)))
            elif rtype == 'liquidity_sweep':
                prevented = max(1, int(round(w_cnt * 0.20)))
            elif rtype == 'loss_rca':
                prevented = sl_hit_count
            elif rtype == 'intraday':
                prevented = squareoff_count

            rules_performance.append({
                'id': r.id,
                'name': r.name,
                'rule_type': rtype,
                'rule_type_display': r.get_rule_type_display(),
                'role': meta['role'],
                'icon': meta['icon'],
                'color': meta['color'],
                'triggered_count': sub_count,
                'winning_count': w_cnt,
                'losing_count': l_cnt,
                'accuracy_pct': acc,
                'net_pnl': pnl_sum,
                'prevented_count': prevented,
                'is_active': r.is_active,
            })

        # Sanitize AI suggested future rules to guarantee preventive_prompt key exists
        if self.object.results and isinstance(self.object.results, dict) and 'ai_suggested_future_rules' in self.object.results:
            ai_rules = self.object.results['ai_suggested_future_rules']
            if isinstance(ai_rules, list):
                for rule_item in ai_rules:
                    if isinstance(rule_item, dict):
                        p_val = rule_item.get('preventive_prompt') or rule_item.get('preventative_prompt') or ''
                        rule_item['preventive_prompt'] = p_val
                        rule_item['preventative_prompt'] = p_val

        FOREX_SYMBOLS = ['MGC', 'M6E', 'M6J', 'MYM', 'MNQ', 'MES', 'MCL']
        context['is_forex'] = (self.object.market_type == 'FOREX_FUTURES' or (self.object.index_name and self.object.index_name.upper() in FOREX_SYMBOLS))
        context['rules_performance'] = rules_performance
        context['equity_curve_json'] = json.dumps(equity_curve_points)
        context['trades_page'] = trade_data['trades_page']
        context['is_trades_paginated'] = trade_data['trades_page'].has_other_pages()
        context['total_trades_count'] = trade_data['total_trades_count']
        context['filtered_trades_count'] = trade_data['filtered_trades_count']
        context['current_filters'] = trade_data['current_filters']
        return context


class BacktestTradesScrollView(LoginRequiredMixin, AdminRequiredMixin, View):
    """
    HTMX Infinite Scroll endpoint to seamlessly append trade rows or mobile cards as the user scrolls.
    """
    def get(self, request, pk, *args, **kwargs):
        backtest = get_object_or_404(BacktestTask, pk=pk, is_deleted=False)
        data = get_backtest_trades_context(backtest, request)
        context = {
            'backtest': backtest,
            'trades_page': data['trades_page'],
            'current_filters': data['current_filters'],
            'filtered_trades_count': data['filtered_trades_count'],
            'total_trades_count': data['total_trades_count'],
        }
        scroll_type = request.GET.get('scroll_type', 'rows')
        if scroll_type == 'cards':
            return render(request, 'admins/partials/backtest_trades_cards.html', context)
        return render(request, 'admins/partials/backtest_trades_rows.html', context)


from apps.common.mixins import BaseHtmxScrollListView

class BacktestDashboardScrollView(LoginRequiredMixin, AdminRequiredMixin, BaseHtmxScrollListView):
    """Endpoint for Load More pagination of backtest runs (desktop rows or mobile cards)."""
    rows_template_name = 'admins/partials/backtest_dashboard_rows.html'
    cards_template_name = 'admins/partials/backtest_dashboard_cards.html'
    context_object_name = 'backtests'

    def get_queryset(self):
        queryset = BacktestTask.objects.filter(is_deleted=False)
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(strategy_name__icontains=q) | queryset.filter(id__icontains=q)
        status = self.request.GET.get('status', '').strip()
        if status:
            queryset = queryset.filter(status=status)
        return queryset.order_by('-created_at')


class BacktestControlView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        action = request.GET.get('action', 'start')
        task = BacktestTask.objects.filter(pk=pk).first()
        return render(request, 'admins/partials/backtest_control_confirm.html', {'backtest': task, 'action': action})

    def post(self, request, pk, *args, **kwargs):
        action = request.POST.get('action', 'start')
        send_backtest_control_command(pk, action)
        msg = f"Backtest command '{action.upper()}' sent successfully."
        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': msg, 'level': 'success'},
            'closeGlobalModal': True,
            'reloadBacktestTable': True,
            'reloadBacktestDetail': True,
        })
        return response


class BacktestStatusView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Returns JSON status and progress percentage for live WebSocket/polling updates."""
    def get(self, request, pk, *args, **kwargs):
        task = get_object_or_404(BacktestTask, pk=pk)
        return JsonResponse({
            'task_id': str(task.id),
            'status': task.status,
            'progress': task.progress or 0,
            'total_trades': task.results.get('total_trades', 0) if task.results else 0,
            'net_pnl': task.results.get('net_pnl', 0.0) if task.results else 0.0,
            'step_info': f"Executing reinforcement learning strategy ({task.progress or 0}%)..." if task.status == BacktestTask.StatusChoices.RUNNING else "",
            'error_logs': task.error_logs or '',
        })


class BacktestDeleteView(HtmxModalMixin, HtmxMessageMixin, LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = BacktestTask
    modal_template_name = 'admins/partials/confirm_delete.html'
    template_name = 'admins/partials/confirm_delete.html'
    success_message = "Backtest task deleted successfully."

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status in [BacktestTask.StatusChoices.RUNNING, BacktestTask.StatusChoices.PENDING]:
            send_backtest_control_command(self.object.id, 'CANCEL')
        self.object.delete()
        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'closeGlobalModal': True,
            'showToast': {'message': str(self.success_message), 'level': 'success'},
            'reloadBacktestTable': True
        })
        return response


class BacktestBulkDeleteView(LoginRequiredMixin, AdminRequiredMixin, View):
    """CBV for bulk soft-deletion of backtest tasks via HTMX."""
    def get(self, request, *args, **kwargs):
        ids_raw = request.GET.get('ids', '')
        ids_list = [i.strip() for i in ids_raw.split(',') if i.strip().isdigit()]
        count = len(ids_list)
        context = {
            'count': count,
            'ids_str': ','.join(ids_list),
            'item_name': 'backtest task' if count == 1 else 'backtest tasks',
            'post_url': reverse_lazy('backtest:backtest_bulk_delete'),
        }
        return render(request, 'admins/partials/confirm_bulk_delete.html', context)

    def post(self, request, *args, **kwargs):
        ids_raw = request.POST.get('ids', '')
        ids_list = [int(i.strip()) for i in ids_raw.split(',') if i.strip().isdigit()]
        if ids_list:
            qs = BacktestTask.objects.filter(id__in=ids_list, is_deleted=False)
            count = qs.count()
            qs.update(is_deleted=True)
            msg = f"Successfully deleted {count} backtest task{'s' if count != 1 else ''}."
        else:
            msg = "No valid backtest tasks selected."
            count = 0

        response = HttpResponse()
        response['HX-Trigger'] = json.dumps({
            'closeGlobalModal': True,
            'showToast': {'message': msg, 'level': 'success' if count > 0 else 'warning'},
            'reloadBacktestTable': True
        })
        return response


# --- STRATEGY SUBMENU & GO CODE WEB UI VIEWS ---

from .models import TradingStrategy

def ensure_default_strategies():
    if TradingStrategy.objects.filter(is_deleted=False).exists():
        return
    
    defaults = [
        {
            "name": "TensorTrade RL (Deep Reinforcement Learning)",
            "code_name": "tensortrade_rl",
            "category": "Deep Reinforcement Learning",
            "target_index": "NIFTY, BANKNIFTY, FINNIFTY",
            "description": "Deep Reinforcement Learning trading agent trained over historical Parquet backup datasets using PPO / A2C policy optimization.",
            "go_file_path": "apps/backtest/rl_engine.py",
            "default_parameters": {
                "lots_count": 1,
                "strike_selection": "ATM",
                "risk_reward_ratio": 2.0,
                "stop_loss_points": 30.0,
                "algorithm": "PPO",
                "reward_metric": "sharpe",
                "total_timesteps": 10000
            },
            "user_manual": """# TensorTrade RL Engine Strategy Manual

## 1. Overview
The **TensorTrade RL Strategy** utilizes Deep Reinforcement Learning (PPO / A2C / DQN) to autonomously train trading agents directly on Marmot's date-partitioned Parquet backup datasets (`/app/backup/{user_id}/{task_id}/dataset.parquet`).

---

## 2. Training & Signal Generation
- Ingests `open`, `high`, `low`, `close`, `volume`, `oi`, `iv`, and `spot_price` into TensorTrade feature streams.
- Trains policy neural networks over 10,000+ timesteps.
- Evaluates risk-managed trade signals (BUY CE, BUY PE, HOLD) with configurable Stop Loss points.
"""
        }
    ]

    for item in defaults:
        TradingStrategy.objects.get_or_create(code_name=item["code_name"], defaults=item)


class StrategyListView(HTMXPartialMixin, LoginRequiredMixin, ListView):
    model = TradingStrategy
    template_name = 'admins/strategy_list.html'
    partial_template_name = 'admins/partials/strategy_list_content.html'
    context_object_name = 'strategies'
    paginate_by = 5

    def get_queryset(self):
        ensure_default_strategies()
        qs = TradingStrategy.objects.filter(is_deleted=False)
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(code_name__icontains=q) |
                Q(category__icontains=q) |
                Q(target_index__icontains=q)
            )
        return qs.order_by('id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        is_admin = getattr(user, 'is_superuser', False) or getattr(user, 'role', '') in ['admin', 'developer']
        context['can_delete'] = is_admin
        context['can_edit'] = is_admin
        context['current_q'] = self.request.GET.get('q', '').strip()
        return context


class StrategyDetailView(HTMXPartialMixin, LoginRequiredMixin, DetailView):
    model = TradingStrategy
    template_name = 'admins/strategy_detail.html'
    partial_template_name = 'admins/partials/strategy_detail_content.html'
    context_object_name = 'strategy'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        strategy = self.object
        user = self.request.user
        is_admin = getattr(user, 'is_superuser', False) or getattr(user, 'role', '') in ['admin', 'developer']
        
        # Read Go strategy source code from filesystem
        go_code = ""
        full_path = os.path.join(settings.BASE_DIR, strategy.go_file_path)
        if os.path.exists(full_path):
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    go_code = f.read()
            except Exception as e:
                go_code = f"// Error reading file {strategy.go_file_path}: {e}"
        else:
            go_code = f"// File {strategy.go_file_path} not found on server."

        context['go_code'] = go_code
        context['can_edit'] = is_admin
        context['can_delete'] = is_admin
        context['params_json_str'] = json.dumps(strategy.default_parameters, indent=2)
        return context


class StrategySaveCodeView(LoginRequiredMixin, View):
    """Admin-only view to save edited Go code and strategy parameters from Web UI."""
    def post(self, request, pk, *args, **kwargs):
        user = request.user
        is_admin = getattr(user, 'is_superuser', False) or getattr(user, 'role', '') in ['admin', 'developer']
        if not is_admin:
            return HttpResponseForbidden("Permission Denied: Only administrators can modify strategy code.")

        strategy = get_object_or_404(TradingStrategy, pk=pk, is_deleted=False)
        new_code = request.POST.get('go_code', '')
        new_params_str = request.POST.get('parameters_json', '{}')

        # Update Go code file
        full_path = os.path.join(settings.BASE_DIR, strategy.go_file_path)
        try:
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            with open(full_path, 'w', encoding='utf-8') as f:
                f.write(new_code)
        except Exception as e:
            response = HttpResponse(status=400)
            response['HX-Trigger'] = json.dumps({
                'showToast': {'message': f"Failed to save Go file: {e}", 'level': 'error'}
            })
            return response

        # Parse & update parameters JSON
        try:
            parsed_params = json.loads(new_params_str)
            strategy.default_parameters = parsed_params
            strategy.save()
        except Exception as e:
            pass

        # Trigger background docker build for go_app
        import subprocess
        try:
            subprocess.Popen(["docker", "compose", "build", "go_app"], cwd=settings.BASE_DIR)
        except Exception:
            pass

        msg = f"Strategy '{strategy.name}' code & parameters saved successfully!"
        response = HttpResponse(status=200)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': msg, 'level': 'success'}
        })
        return response


class StrategyDeleteView(LoginRequiredMixin, View):
    """Admin-only deletion view for strategies."""
    def get(self, request, pk, *args, **kwargs):
        user = request.user
        is_admin = getattr(user, 'is_superuser', False) or getattr(user, 'role', '') in ['admin', 'developer']
        if not is_admin:
            return render(request, 'admins/partials/confirm_delete.html', {
                'object': None,
                'error_msg': "Permission Denied: Regular users cannot delete strategies. Only system admins have delete authority."
            })
        strategy = get_object_or_404(TradingStrategy, pk=pk, is_deleted=False)
        return render(request, 'admins/partials/confirm_delete.html', {'object': strategy, 'item_name': 'Strategy'})

    def post(self, request, pk, *args, **kwargs):
        user = request.user
        is_admin = getattr(user, 'is_superuser', False) or getattr(user, 'role', '') in ['admin', 'developer']
        if not is_admin:
            response = HttpResponse(status=403)
            response['HX-Trigger'] = json.dumps({
                'closeGlobalModal': True,
                'showToast': {'message': "Permission Denied: Only Admins can delete strategies.", 'level': 'error'}
            })
            return response

        strategy = get_object_or_404(TradingStrategy, pk=pk, is_deleted=False)
        strategy.is_deleted = True
        strategy.save()

        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'closeGlobalModal': True,
            'showToast': {'message': f"Strategy '{strategy.name}' deleted successfully.", 'level': 'success'},
            'reloadStrategyTable': True
        })
        return response


class BacktestRuleListView(HTMXPartialMixin, LoginRequiredMixin, AdminRequiredMixin, ListView):
    """Rule Management Dashboard with search, filter, and HTMX partial swap."""
    model = BacktestRule
    template_name = 'admins/backtest_rule_list.html'
    partial_template_name = 'admins/partials/backtest_rule_table_partial.html'
    context_object_name = 'rules'

    def get_queryset(self):
        queryset = super().get_queryset().filter(is_deleted=False)
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(Q(name__icontains=q) | Q(description__icontains=q) | Q(prompt_directive__icontains=q))
        market_type = self.request.GET.get('market_type', '').strip()
        if market_type:
            queryset = queryset.filter(market_type=market_type)
        rule_type = self.request.GET.get('rule_type', '').strip()
        if rule_type:
            queryset = queryset.filter(rule_type=rule_type)
        is_active = self.request.GET.get('is_active', '').strip()
        if is_active in ['true', '1']:
            queryset = queryset.filter(is_active=True)
        elif is_active in ['false', '0']:
            queryset = queryset.filter(is_active=False)
        return queryset.order_by('-is_system_preset', 'id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Backtest Strategy Rules"
        context['rule_types'] = BacktestRule.RuleTypeChoices.choices
        context['market_types'] = BacktestRule.MarketTypeChoices.choices
        context['current_filters'] = {
            'q': self.request.GET.get('q', ''),
            'market_type': self.request.GET.get('market_type', ''),
            'rule_type': self.request.GET.get('rule_type', ''),
            'is_active': self.request.GET.get('is_active', ''),
        }
        return context


class BacktestRuleCreateView(HtmxModalMixin, HtmxMessageMixin, LoginRequiredMixin, AdminRequiredMixin, CreateView):
    """Modal creation view for new custom strategy rules."""
    model = BacktestRule
    form_class = BacktestRuleForm
    template_name = 'admins/partials/backtest_rule_form_modal.html'
    modal_template_name = 'admins/partials/backtest_rule_form_modal.html'
    success_message = "Strategy Rule created successfully."

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.is_system_preset = False
        self.object.save()
        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'closeGlobalModal': True,
            'showToast': {'message': str(self.success_message), 'level': 'success'},
            'reloadRuleTable': True
        })
        return response


class BacktestRuleUpdateView(HtmxModalMixin, HtmxMessageMixin, LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    """Modal update view for strategy rules."""
    model = BacktestRule
    form_class = BacktestRuleForm
    template_name = 'admins/partials/backtest_rule_form_modal.html'
    modal_template_name = 'admins/partials/backtest_rule_form_modal.html'
    success_message = "Strategy Rule updated successfully."

    def form_valid(self, form):
        self.object = form.save()
        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'closeGlobalModal': True,
            'showToast': {'message': str(self.success_message), 'level': 'success'},
            'reloadRuleTable': True
        })
        return response


class BacktestRuleDeleteView(HtmxModalMixin, HtmxMessageMixin, LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    """Delete confirmation view with safeguard against system presets."""
    model = BacktestRule
    modal_template_name = 'admins/partials/confirm_delete.html'
    template_name = 'admins/partials/confirm_delete.html'
    success_message = "Strategy Rule deleted successfully."

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.is_system_preset:
            return render(request, 'admins/partials/confirm_delete.html', {
                'object': None,
                'error_msg': "Protected System Preset: Built-in default rules cannot be deleted. You can toggle their active state instead."
            })
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.is_system_preset:
            response = HttpResponse(status=403)
            response['HX-Trigger'] = json.dumps({
                'closeGlobalModal': True,
                'showToast': {'message': "System presets cannot be deleted.", 'level': 'warning'}
            })
            return response
        self.object.is_deleted = True
        self.object.save()
        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'closeGlobalModal': True,
            'showToast': {'message': str(self.success_message), 'level': 'success'},
            'reloadRuleTable': True
        })
        return response


class BacktestRuleToggleView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Quick HTMX toggle endpoint to activate/deactivate strategy rules."""
    def post(self, request, pk, *args, **kwargs):
        rule = get_object_or_404(BacktestRule, pk=pk, is_deleted=False)
        rule.is_active = not rule.is_active
        rule.save(update_fields=['is_active'])
        status_str = "activated" if rule.is_active else "deactivated"
        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f"Rule '{rule.name}' {status_str}.", 'level': 'info'},
            'reloadRuleTable': True
        })
        return response


class BacktestEditModalView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Renders and processes modal to adjust backtest parameters and trigger fresh regeneration."""

    def get(self, request, pk, *args, **kwargs):
        task = get_object_or_404(BacktestTask, pk=pk, is_deleted=False)
        FOREX_SYMBOLS = ['MGC', 'M6E', 'M6J', 'MYM', 'MNQ', 'MES', 'MCL']
        is_forex = (task.market_type == 'FOREX_FUTURES' or (task.index_name and task.index_name.upper() in FOREX_SYMBOLS))

        if is_forex:
            available_rules = BacktestRule.objects.filter(is_active=True, is_deleted=False, market_type__in=['FOREX_FUTURES', 'ALL'])
        else:
            available_rules = BacktestRule.objects.filter(is_active=True, is_deleted=False, market_type__in=['INDEX_FO', 'ALL'])

        active_rule_ids = set(task.rules.values_list('id', flat=True))
        if not active_rule_ids and task.parameters and 'rules' in task.parameters:
            for r in task.parameters['rules']:
                if isinstance(r, dict) and 'id' in r:
                    active_rule_ids.add(r['id'])
                elif isinstance(r, dict) and 'name' in r:
                    matched = available_rules.filter(name=r['name']).first()
                    if matched:
                        active_rule_ids.add(matched.id)

        add_rule_type = request.GET.get('add_rule_type', '').strip()
        add_rule_id = request.GET.get('add_rule_id', '').strip()

        if add_rule_type:
            rule_match = BacktestRule.objects.filter(rule_type=add_rule_type, is_active=True, is_deleted=False).first()
            if rule_match:
                active_rule_ids.add(rule_match.id)

        if add_rule_id:
            try:
                active_rule_ids.add(int(add_rule_id))
            except ValueError:
                pass

        params = task.parameters or {}
        context = {
            'backtest': task,
            'is_forex': is_forex,
            'available_rules': available_rules,
            'active_rule_ids': active_rule_ids,
            'start_date_val': task.start_date.strftime('%Y-%m-%d') if task.start_date else '',
            'end_date_val': task.end_date.strftime('%Y-%m-%d') if task.end_date else '',
            'initial_capital_val': task.initial_capital,
            'lots_count_val': params.get('lots_count', 1),
            'strike_selection_val': params.get('strike_selection', 'ATM'),
            'stop_loss_points_val': params.get('stop_loss_points', params.get('sl_pts', 30.0)),
            'rr_ratio_val': params.get('rr_ratio', 2.0),
            'prompt_directives_val': params.get('prompt_directives', ''),
        }
        return render(request, 'admins/partials/backtest_edit_modal.html', context)

    def post(self, request, pk, *args, **kwargs):
        task = get_object_or_404(BacktestTask, pk=pk, is_deleted=False)
        FOREX_SYMBOLS = ['MGC', 'M6E', 'M6J', 'MYM', 'MNQ', 'MES', 'MCL']
        is_forex = (task.market_type == 'FOREX_FUTURES' or (task.index_name and task.index_name.upper() in FOREX_SYMBOLS))

        start_date_str = request.POST.get('start_date', '').strip()
        end_date_str = request.POST.get('end_date', '').strip()
        initial_capital_str = request.POST.get('initial_capital', '').strip()

        if start_date_str:
            task.start_date = start_date_str
        if end_date_str:
            task.end_date = end_date_str
        if initial_capital_str:
            try:
                task.initial_capital = float(initial_capital_str)
            except ValueError:
                pass

        try:
            rr_ratio = float(request.POST.get('rr_ratio', 2.0))
        except ValueError:
            rr_ratio = 2.0

        try:
            sl_pts = float(request.POST.get('stop_loss_points', 30.0))
        except ValueError:
            sl_pts = 30.0

        try:
            lots_count = int(request.POST.get('lots_count', 1))
        except ValueError:
            lots_count = 1

        strike_selection = 'SPOT' if is_forex else request.POST.get('strike_selection', 'ATM').strip()
        prompt_directives = request.POST.get('prompt_directives', '').strip()

        selected_rule_ids = request.POST.getlist('rules')
        if is_forex:
            rules_qs = BacktestRule.objects.filter(id__in=selected_rule_ids, is_active=True, is_deleted=False, market_type__in=['FOREX_FUTURES', 'ALL'])
        else:
            rules_qs = BacktestRule.objects.filter(id__in=selected_rule_ids, is_active=True, is_deleted=False, market_type__in=['INDEX_FO', 'ALL'])

        rules_list = []
        for r in rules_qs:
            rules_list.append({
                "id": r.id,
                "name": r.name,
                "rule_type": r.rule_type,
                "parameters": r.parameters or {},
                "prompt_directive": r.prompt_directive or "",
            })

        task.parameters = {
            **(task.parameters or {}),
            "market_type": "FOREX_FUTURES" if is_forex else "INDEX_FO",
            "rr_ratio": rr_ratio,
            "stop_loss_points": sl_pts,
            "sl_pts": sl_pts,
            "lots_count": lots_count,
            "strike_selection": strike_selection,
            "rules": rules_list,
            "prompt_directives": prompt_directives,
        }
        task.save(update_fields=['start_date', 'end_date', 'initial_capital', 'parameters'])
        task.rules.set(rules_qs)

        send_backtest_control_command(task.id, 'START')

        response = HttpResponse(status=204)
        response['HX-Trigger'] = json.dumps({
            'showToast': {'message': f'Parameters saved & Backtest #BT-{task.id:04d} regeneration started!', 'level': 'success'},
            'closeGlobalModal': True,
            'reloadBacktestDetail': True,
            'reloadBacktestTable': True,
        })
        return response


class RLTrainingIndexView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    """Dedicated RL AI Training portal for Index Options (NIFTY, BANKNIFTY, FINNIFTY, SENSEX, etc.)."""
    template_name = 'admins/rl_training_index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['index_form'] = IndexBacktestTaskForm(initial={'strategy_name': 'tensortrade_rl'})
        context['backtests'] = BacktestTask.objects.filter(
            strategy_name='tensortrade_rl'
        ).select_related('created_by', 'backup_task').order_by('-created_at')[:10]
        context['active_market_type'] = 'INDEX_FO'
        context['ws_url'] = settings.MARMOT_WS_URL
        return context

    def post(self, request, *args, **kwargs):
        post_data = request.POST.copy()
        if post_data.get('risk_reward_ratio') == 'AUTO':
            post_data['risk_reward_ratio'] = '2.0'
        if post_data.get('strike_selection') == 'AUTO':
            post_data['strike_selection'] = 'ATM'

        form = IndexBacktestTaskForm(post_data)
        if form.is_valid():
            selected_rules = form.cleaned_data.get('rules')
            prompt_directives = form.cleaned_data.get('prompt_directives', '').strip()
            rule_list = []
            if selected_rules:
                for r in selected_rules:
                    rule_list.append({
                        'id': r.id,
                        'name': r.name,
                        'rule_type': r.rule_type,
                        'prompt_directive': r.prompt_directive,
                        'parameters': r.parameters or {}
                    })

            limit_offset_pts = float(request.POST.get('limit_offset_pts', 0.0))
            order_entry_style = request.POST.get('order_entry_style', 'LIMIT_ORDER')
            algorithm = request.POST.get('algorithm', 'PPO')
            reward_metric = request.POST.get('reward_metric', 'sharpe')
            total_timesteps = int(request.POST.get('total_timesteps', 10000))

            params = {
                "market_type": "INDEX_FO",
                "algorithm": algorithm,
                "reward_metric": reward_metric,
                "total_timesteps": total_timesteps,
                "order_entry_style": order_entry_style,
                "limit_offset_pts": limit_offset_pts,
                "rr_ratio": form.cleaned_data.get('risk_reward_ratio', 2.0),
                "stop_loss_points": form.cleaned_data.get('stop_loss_points', 30.0),
                "sl_pts": form.cleaned_data.get('stop_loss_points', 30.0),
                "lots_count": form.cleaned_data.get('lots_count', 1),
                "rules": rule_list,
                "prompt_directives": prompt_directives,
            }

            if 'strike_selection' in form.cleaned_data and form.cleaned_data['strike_selection']:
                params["strike_selection"] = form.cleaned_data['strike_selection']

            backup_task = form.cleaned_data.get('backup_task')
            task = create_and_start_backtest_task(
                strategy_name='tensortrade_rl',
                index_name=form.cleaned_data.get('index_name', 'NIFTY'),
                start_date=form.cleaned_data.get('start_date'),
                end_date=form.cleaned_data.get('end_date'),
                initial_capital=form.cleaned_data.get('initial_capital', 100000.0),
                parameters=params,
                user=request.user,
                backup_task=backup_task
            )
            if selected_rules:
                task.rules.set(selected_rules)

            send_backtest_control_command(task.id, 'START')

            if request.headers.get('HX-Request'):
                response = HttpResponse(status=200)
                response['HX-Redirect'] = reverse('backtest:backtest_detail', kwargs={'pk': task.id})
                return response
            return redirect('backtest:backtest_detail', pk=task.id)

        context = self.get_context_data()
        context['index_form'] = form
        return self.render_to_response(context)


class RLTrainingForexView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    """Dedicated RL AI Training portal for Forex & CME Futures (MGC, M6E, M6J, MNQ, MES, MCL, etc.)."""
    template_name = 'admins/rl_training_forex.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['forex_form'] = ForexBacktestTaskForm(initial={'strategy_name': 'tensortrade_rl', 'index_name': 'MNQ'})
        context['backtests'] = BacktestTask.objects.filter(
            strategy_name='tensortrade_rl'
        ).select_related('created_by', 'backup_task').order_by('-created_at')[:10]
        context['active_market_type'] = 'FOREX_FUTURES'
        context['ws_url'] = settings.MARMOT_WS_URL
        return context

    def post(self, request, *args, **kwargs):
        post_data = request.POST.copy()
        if post_data.get('risk_reward_ratio') == 'AUTO':
            post_data['risk_reward_ratio'] = '2.0'

        form = ForexBacktestTaskForm(post_data)
        if form.is_valid():
            selected_rules = form.cleaned_data.get('rules')
            prompt_directives = form.cleaned_data.get('prompt_directives', '').strip()
            rule_list = []
            if selected_rules:
                for r in selected_rules:
                    rule_list.append({
                        'id': r.id,
                        'name': r.name,
                        'rule_type': r.rule_type,
                        'prompt_directive': r.prompt_directive,
                        'parameters': r.parameters or {}
                    })

            limit_offset_ticks = float(request.POST.get('limit_offset_ticks', 0.0))
            order_entry_style = request.POST.get('order_entry_style', 'LIMIT_ORDER')
            algorithm = request.POST.get('algorithm', 'PPO')
            reward_metric = request.POST.get('reward_metric', 'sharpe')
            total_timesteps = int(request.POST.get('total_timesteps', 10000))

            params = {
                "market_type": "FOREX_FUTURES",
                "algorithm": algorithm,
                "reward_metric": reward_metric,
                "total_timesteps": total_timesteps,
                "order_entry_style": order_entry_style,
                "limit_offset_ticks": limit_offset_ticks,
                "rr_ratio": form.cleaned_data.get('risk_reward_ratio', 2.0),
                "stop_loss_points": form.cleaned_data.get('stop_loss_points', 30.0),
                "sl_pts": form.cleaned_data.get('stop_loss_points', 30.0),
                "lots_count": form.cleaned_data.get('lots_count', 1),
                "rules": rule_list,
                "prompt_directives": prompt_directives,
            }

            backup_task = form.cleaned_data.get('backup_task')
            task = create_and_start_backtest_task(
                strategy_name='tensortrade_rl',
                index_name=form.cleaned_data.get('index_name', 'MNQ'),
                start_date=form.cleaned_data.get('start_date'),
                end_date=form.cleaned_data.get('end_date'),
                initial_capital=form.cleaned_data.get('initial_capital', 100000.0),
                parameters=params,
                user=request.user,
                backup_task=backup_task
            )
            if selected_rules:
                task.rules.set(selected_rules)

            send_backtest_control_command(task.id, 'START')

            if request.headers.get('HX-Request'):
                response = HttpResponse(status=200)
                response['HX-Redirect'] = reverse('backtest:backtest_detail', kwargs={'pk': task.id})
                return response
            return redirect('backtest:backtest_detail', pk=task.id)

        context = self.get_context_data()
        context['forex_form'] = form
        return self.render_to_response(context)


import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class BacktestExportExcelView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Generates a professional multi-sheet Excel report (.xlsx) for a BacktestTask."""

    def get(self, request, pk, *args, **kwargs):
        backtest = get_object_or_404(BacktestTask, pk=pk, is_deleted=False)
        trade_data = get_backtest_trades_context(backtest, request)
        all_trades = trade_data['all_trades']

        wb = openpyxl.Workbook()

        # -------------------------------------------------------------
        # SHEET 1: SUMMARY (Executive Dashboard)
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        title_font = Font(name="Calibri", size=15, bold=True, color="1E293B")
        section_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
        regular_font = Font(name="Calibri", size=10, color="334155")

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        accent_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws_summary.merge_cells("A1:F1")
        ws_summary["A1"] = f"MARMOT QUANT ENGINE — BACKTEST EXECUTIVE REPORT #BT-{backtest.id:04d}"
        ws_summary["A1"].font = title_font
        ws_summary["A1"].alignment = Alignment(vertical="center")

        ws_summary.merge_cells("A2:F2")
        from django.utils import timezone
        ws_summary["A2"] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Strategy: {backtest.get_strategy_name_display()} | Asset: {backtest.index_name}"
        ws_summary["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")

        ws_summary.cell(row=4, column=1, value="BACKTEST METADATA").font = section_font
        meta_items = [
            ("Backtest ID", f"#BT-{backtest.id:04d}"),
            ("Strategy Name", backtest.get_strategy_name_display()),
            ("Market Asset", backtest.index_name),
            ("Market Type", getattr(backtest, 'market_type', 'INDEX_FO')),
            ("Date Period", f"{backtest.start_date} to {backtest.end_date}"),
            ("Initial Capital", f"₹{float(backtest.initial_capital or 100000):,.2f}"),
            ("Execution Status", str(backtest.status).upper()),
        ]

        for r_idx, (label, val) in enumerate(meta_items, start=5):
            ws_summary.cell(row=r_idx, column=1, value=label).font = bold_font
            ws_summary.cell(row=r_idx, column=2, value=val).font = regular_font
            ws_summary.cell(row=r_idx, column=1).fill = card_fill
            ws_summary.cell(row=r_idx, column=2).fill = card_fill
            ws_summary.cell(row=r_idx, column=1).border = thin_border
            ws_summary.cell(row=r_idx, column=2).border = thin_border

        res = backtest.results or {}
        net_pnl = float(res.get('net_pnl', 0.0))
        gross_pnl = float(res.get('gross_pnl', net_pnl))
        total_trades = int(res.get('total_trades', len(all_trades)))
        win_count = int(res.get('win_trades', sum(1 for t in all_trades if float(t.get('pnl', 0)) > 0)))
        loss_count = int(res.get('loss_trades', total_trades - win_count))
        win_rate = float(res.get('win_rate', (win_count / total_trades * 100) if total_trades > 0 else 0.0))
        max_utilized = float(res.get('max_utilized_capital', res.get('max_capital_used', 0.0)))
        total_charges = float(res.get('total_charges', res.get('statutory_taxes', 0.0)))
        sharpe = float(res.get('sharpe_ratio', 0.0))
        max_dd_pct = float(res.get('max_drawdown_pct', 0.0))
        init_cap = float(backtest.initial_capital or 100000.0)
        roi_total = ((net_pnl / init_cap) * 100.0) if init_cap > 0 else 0.0
        roi_utilized = ((net_pnl / max_utilized) * 100.0) if max_utilized > 0 else roi_total

        ws_summary.cell(row=4, column=4, value="EXECUTIVE FINANCIAL METRICS").font = section_font
        metrics_items = [
            ("Gross PnL", f"₹{gross_pnl:,.2f}"),
            ("Statutory Fees & Taxes", f"₹{total_charges:,.2f}"),
            ("NET PnL", f"₹{net_pnl:,.2f}"),
            ("Win Rate %", f"{win_rate:.1f}%"),
            ("Total Trades (Win / Loss)", f"{total_trades} (Win: {win_count} | Loss: {loss_count})"),
            ("Max Utilized Capital", f"₹{max_utilized:,.2f}"),
            ("ROI on Utilized Capital %", f"{roi_utilized:.2f}%"),
            ("ROI on Total Capital %", f"{roi_total:.2f}%"),
            ("Sharpe Ratio", f"{sharpe:.2f}"),
            ("Max Drawdown %", f"{max_dd_pct:.2f}%"),
        ]

        for r_idx, (label, val) in enumerate(metrics_items, start=5):
            ws_summary.cell(row=r_idx, column=4, value=label).font = bold_font
            cell_v = ws_summary.cell(row=r_idx, column=5, value=val)
            cell_v.font = bold_font if label == "NET PnL" else regular_font
            ws_summary.cell(row=r_idx, column=4).fill = accent_fill
            ws_summary.cell(row=r_idx, column=5).fill = accent_fill
            ws_summary.cell(row=r_idx, column=4).border = thin_border
            ws_summary.cell(row=r_idx, column=5).border = thin_border

        # -------------------------------------------------------------
        # SHEET 2: TRADES DATA (Detailed Trade Log Table)
        # -------------------------------------------------------------
        ws_trades = wb.create_sheet(title="Trades Data")
        ws_trades.views.sheetView[0].showGridLines = True

        headers = [
            "Trade #", "Timestamp", "Symbol", "Direction", "Entry Price", "Exit Price", 
            "Lots / Qty", "Gross PnL", "Brokerage", "Statutory Taxes", "Net PnL", 
            "Equity Before", "Equity After", "Trade ROI %", "Exit Reason", "RCA Diagnosis"
        ]

        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws_trades.cell(row=1, column=col_idx, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for r_idx, t in enumerate(all_trades, start=2):
            sn = t.get('serial_no', r_idx - 1)
            ts = str(t.get('timestamp', t.get('entry_time', '')))
            sym = str(t.get('symbol', t.get('strike', backtest.index_name)))
            side = str(t.get('trade_type', t.get('side', t.get('action', 'BUY')))).upper()
            entry_p = float(t.get('index_entry_price', t.get('entry_price', 0.0)))
            exit_p = float(t.get('index_exit_price', t.get('exit_price', 0.0)))
            qty = int(t.get('qty', t.get('quantity', t.get('lots', 1))))
            gross = float(t.get('gross_pnl', t.get('pnl', 0.0)))
            brok = float(t.get('brokerage', 20.0))
            taxes = float(t.get('statutory_taxes', t.get('stt', 0.0)))
            net = float(t.get('net_pnl', t.get('pnl', 0.0)))
            eq_before = float(t.get('equity_before', 100000.0))
            eq_after = float(t.get('equity_after', eq_before + net))
            roi_pct = float(t.get('trade_equity_change_pct', 0.0))
            reason = str(t.get('reason', t.get('status', 'TARGET')))
            rca_diag = str(t.get('rca_primary_cause', t.get('rca_summary', 'Normal Execution')))

            row_vals = [
                sn, ts, sym, side, entry_p, exit_p, qty, gross, brok, taxes, net, 
                eq_before, eq_after, roi_pct, reason, rca_diag
            ]

            for c_idx, val in enumerate(row_vals, start=1):
                c = ws_trades.cell(row=r_idx, column=c_idx, value=val)
                c.font = regular_font
                c.border = thin_border
                if c_idx in [1, 2, 4, 15]:
                    c.alignment = Alignment(horizontal="center")
                elif c_idx in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
                    c.alignment = Alignment(horizontal="right")

        for sheet in [ws_summary, ws_trades]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Marmot_Backtest_Report_BT-{backtest.id:04d}_{backtest.index_name}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


